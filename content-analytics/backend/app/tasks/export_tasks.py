"""导出任务"""
import os
import uuid
import asyncio
from datetime import datetime
from typing import Optional

from sqlalchemy import select
from sqlalchemy.orm import selectinload
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import httpx
from io import BytesIO
from PIL import Image as PILImage

from app.tasks.celery_app import celery_app
from app.db.session import async_session_maker, create_thread_session_maker
from app.models.export import Export, ExportStatus, ExportFormat
from app.models.analysis import Analysis, AnalysisResult, AIOutput
from app.core.config import settings


def run_async(coro):
    """在同步环境中运行异步代码"""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()


@celery_app.task(bind=True, name="run_export_task")
def run_export_task(self, export_id: str):
    """Celery导出任务入口"""
    return run_async(_run_export(export_id, use_thread_session=True))


async def _run_export(export_id: str, use_thread_session: bool = False):
    """执行导出任务"""
    if use_thread_session:
        session_maker, thread_engine = create_thread_session_maker()
    else:
        session_maker = async_session_maker
        thread_engine = None

    try:
        async with session_maker() as db:
            # 获取导出记录
            result = await db.execute(
                select(Export).where(Export.id == uuid.UUID(export_id))
            )
            export = result.scalar_one_or_none()
            
            if not export:
                return {"error": "导出记录不存在"}
            
            # 更新状态为处理中
            export.status = ExportStatus.PROCESSING
            await db.commit()
            
            try:
                # 获取分析数据
                result = await db.execute(
                    select(Analysis)
                    .options(selectinload(Analysis.dataset))
                    .where(Analysis.id == export.analysis_id)
                )
                analysis = result.scalar_one_or_none()
                
                if not analysis:
                    raise Exception("分析任务不存在")
                
                # 获取分析结果（按原始数据集顺序排序）
                from app.models.post import Post
                results = await db.execute(
                    select(AnalysisResult)
                    .options(
                        selectinload(AnalysisResult.post),
                        selectinload(AnalysisResult.ai_output)
                    )
                    .join(Post, AnalysisResult.post_id == Post.id)
                    .where(AnalysisResult.analysis_id == analysis.id)
                    .order_by(Post.created_at.asc())
                )
                analysis_results = results.scalars().all()
                
                # 根据格式导出
                if export.format == ExportFormat.EXCEL:
                    file_path = await _export_to_excel(analysis, analysis_results, export_id)
                elif export.format == ExportFormat.JSON:
                    file_path = await _export_to_json(analysis, analysis_results, export_id)
                else:
                    format_value = getattr(export.format, "value", export.format)
                    raise Exception(f"暂不支持 {format_value} 格式导出")
                
                # 更新导出记录
                export.file_path = file_path
                export.status = ExportStatus.COMPLETED
                export.completed_at = datetime.utcnow()
                await db.commit()
                
                return {"success": True, "file_path": file_path}
                
            except Exception as e:
                export.status = ExportStatus.FAILED
                export.error_message = str(e)[:1000]
                await db.commit()
                return {"error": str(e)}
    finally:
        if thread_engine:
            await thread_engine.dispose()


async def _export_to_excel(analysis: Analysis, results: list, export_id: str) -> str:
    """导出为Excel格式"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分析报告"
    
    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="FF2442", end_color="FF2442", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入标题行（与笔记详情页面一致）
    headers = [
        "序号",
        "笔记ID",
        "笔记链接",
        "发文时间",
        "内容形式",
        "发文类型",
        "素材来源",
        "款式信息",
        "标题",
        "正文",
        "封面图片",
        "表现",
        "问题指标",
        "亮点指标",
        "7天阅读",
        "14天阅读",
        "7天互动",
        "14天互动",
        "7天好物访问",
        "14天好物访问",
        "7天好物想要",
        "14天好物想要",
        "AI一句话总结",
        "优点",
        "问题",
        "优化建议"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row_idx, ar in enumerate(results, 2):
        post = getattr(ar, "post", None)
        
        # 序号
        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
        # 笔记ID
        ws.cell(row=row_idx, column=2, value=(post.data_id if post else "")).border = thin_border
        # 笔记链接
        ws.cell(row=row_idx, column=3, value=(post.publish_link if post else "")).border = thin_border
        # 发文时间
        ws.cell(row=row_idx, column=4, value=(str(post.publish_time) if post and post.publish_time else "")).border = thin_border
        # 内容形式
        ws.cell(row=row_idx, column=5, value=(post.content_type if post else "")).border = thin_border
        # 发文类型
        ws.cell(row=row_idx, column=6, value=(post.post_type if post else "")).border = thin_border
        # 素材来源
        ws.cell(row=row_idx, column=7, value=(post.source if post else "")).border = thin_border
        # 款式信息
        ws.cell(row=row_idx, column=8, value=(post.style_info if post else "")).border = thin_border
        # 标题
        ws.cell(row=row_idx, column=9, value=(post.content_title if post else "")).border = thin_border
        # 正文
        ws.cell(row=row_idx, column=10, value=(post.content_text if post else "")).border = thin_border
        
        # 封面图片（嵌入缩略图）
        ws.cell(row=row_idx, column=11, value="").border = thin_border
        if post and post.cover_image:
            try:
                # 下载图片
                with httpx.Client(timeout=10.0) as client:
                    resp = client.get(post.cover_image)
                    if resp.status_code == 200:
                        img_data = BytesIO(resp.content)
                        pil_img = PILImage.open(img_data)
                        # 转换为RGB（处理webp等格式）
                        if pil_img.mode in ('RGBA', 'P'):
                            pil_img = pil_img.convert('RGB')
                        # 缩放到合适大小（宽度200像素，更清晰）
                        ratio = 200 / pil_img.width
                        new_size = (200, int(pil_img.height * ratio))
                        pil_img = pil_img.resize(new_size, PILImage.Resampling.LANCZOS)
                        # 保存到内存
                        img_buffer = BytesIO()
                        pil_img.save(img_buffer, format='PNG')
                        img_buffer.seek(0)
                        # 插入到Excel
                        xl_img = XLImage(img_buffer)
                        ws.add_image(xl_img, f"K{row_idx}")
                        # 设置行高以适应图片
                        ws.row_dimensions[row_idx].height = new_size[1] * 0.75
            except Exception as e:
                # 图片下载失败，写入URL
                ws.cell(row=row_idx, column=11, value=post.cover_image)
        
        # 表现、问题指标、亮点指标
        result_data = ar.result_data or {}
        ws.cell(row=row_idx, column=12, value=(ar.performance or "")).border = thin_border
        ws.cell(row=row_idx, column=13, value="、".join(result_data.get('problem_metrics', []))).border = thin_border
        ws.cell(row=row_idx, column=14, value="、".join(result_data.get('highlight_metrics', []))).border = thin_border
        
        # 数据指标
        ws.cell(row=row_idx, column=15, value=(post.read_7d if post else None)).border = thin_border
        ws.cell(row=row_idx, column=16, value=(post.read_14d if post else None)).border = thin_border
        ws.cell(row=row_idx, column=17, value=(post.interact_7d if post else None)).border = thin_border
        ws.cell(row=row_idx, column=18, value=(post.interact_14d if post else None)).border = thin_border
        ws.cell(row=row_idx, column=19, value=(post.visit_7d if post else None)).border = thin_border
        ws.cell(row=row_idx, column=20, value=(post.visit_14d if post else None)).border = thin_border
        ws.cell(row=row_idx, column=21, value=(post.want_7d if post else None)).border = thin_border
        ws.cell(row=row_idx, column=22, value=(post.want_14d if post else None)).border = thin_border
        
        # AI分析结果
        if ar.ai_output:
            ws.cell(row=row_idx, column=23, value=ar.ai_output.summary or "").border = thin_border
            ws.cell(row=row_idx, column=24, value="\n".join(ar.ai_output.strengths or [])).border = thin_border
            ws.cell(row=row_idx, column=25, value="\n".join(ar.ai_output.weaknesses or [])).border = thin_border
            ws.cell(row=row_idx, column=26, value="\n".join(ar.ai_output.suggestions or [])).border = thin_border
        else:
            for col in range(23, 27):
                ws.cell(row=row_idx, column=col, value="").border = thin_border
        
        # 设置对齐方式
        for col in range(1, 27):
            ws.cell(row=row_idx, column=col).alignment = cell_alignment
    
    # 设置列宽（封面图片列加宽到30适应200px图片，新增表现/问题指标/亮点指标列）
    column_widths = [6, 24, 40, 18, 10, 10, 12, 20, 40, 50, 30, 10, 25, 25, 10, 10, 10, 10, 12, 12, 12, 12, 40, 30, 30, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 保存文件
    export_dir = os.path.join(settings.UPLOAD_DIR, "exports")
    os.makedirs(export_dir, exist_ok=True)
    
    filename = f"analysis_report_{export_id}.xlsx"
    file_path = os.path.join(export_dir, filename)
    wb.save(file_path)
    
    return file_path


async def _export_to_json(analysis: Analysis, results: list, export_id: str) -> str:
    """导出为JSON格式"""
    import json
    
    data = {
        "analysis": {
            "id": str(analysis.id),
            "name": analysis.name,
            "created_at": str(analysis.created_at),
            "status": getattr(analysis.status, "value", analysis.status)
        },
        "results": []
    }
    
    for ar in results:
        post = getattr(ar, "post", None)
        item = {
            "analysis_result": {
                "id": str(ar.id),
                "performance": ar.performance,
                "result_data": ar.result_data,
                "created_at": str(ar.created_at)
            },
            "post": {
                "id": str(post.id) if post else None,
                "data_id": post.data_id if post else None,
                "publish_time": str(post.publish_time) if post and post.publish_time else None,
                "publish_link": post.publish_link if post else None,
                "content_type": post.content_type if post else None,
                "post_type": post.post_type if post else None,
                "source": post.source if post else None,
                "style_info": post.style_info if post else None,
                "metrics": {
                    "read_7d": post.read_7d if post else None,
                    "interact_7d": post.interact_7d if post else None,
                    "visit_7d": post.visit_7d if post else None,
                    "want_7d": post.want_7d if post else None,
                    "read_14d": post.read_14d if post else None,
                    "interact_14d": post.interact_14d if post else None,
                    "visit_14d": post.visit_14d if post else None,
                    "want_14d": post.want_14d if post else None
                }
            },
            "ai_analysis": None
        }

        if ar.ai_output:
            item["ai_analysis"] = {
                "summary": ar.ai_output.summary,
                "strengths": ar.ai_output.strengths,
                "weaknesses": ar.ai_output.weaknesses,
                "suggestions": ar.ai_output.suggestions,
                "model_name": ar.ai_output.model_name,
                "created_at": str(ar.ai_output.created_at)
            }

        data["results"].append(item)
    
    # 保存文件
    export_dir = os.path.join(settings.UPLOAD_DIR, "exports")
    os.makedirs(export_dir, exist_ok=True)
    
    filename = f"analysis_report_{export_id}.json"
    file_path = os.path.join(export_dir, filename)
    
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    return file_path
